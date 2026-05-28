"""
reports/excel_reporter.py

Generates the Excel workbook output.
Replaces ExportToExcel.cs (OpenXML) with openpyxl.

Output: yyyy-MM-dd_OptionVolumeLeaders.xlsx  with four sheets:

  Sheet 1: Stocks           — raw filtered contracts (one row per contract)
  Sheet 2: Stocks Bets      — aggregated bias per ticker  ← main signal sheet
  Sheet 3: ETFs             — same as Sheet 1 for ETFs
  Sheet 4: ETFs Bets        — same as Sheet 2 for ETFs

The "Bets" sheets are colour-coded:
  Long bias   → green fill
  Short bias  → red fill
  Neutral     → no fill
Signal strength column makes the multi-day persistent tickers jump out.
"""
from __future__ import annotations

import logging
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.options import OptionContract, TickerBias

logger = logging.getLogger(__name__)

# Colour fills
_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_RED    = PatternFill("solid", fgColor="FFC7CE")
_GOLD   = PatternFill("solid", fgColor="FFEB9C")   # strong signal highlight
_HEADER = PatternFill("solid", fgColor="2F4F8F")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


class ExcelReporter:

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir

    def write(
        self,
        stock_contracts:  list[OptionContract],
        stock_biases:     list[TickerBias],
        etf_contracts:    list[OptionContract],
        etf_biases:       list[TickerBias],
        report_date: date = None,
    ) -> Path:
        report_date = report_date or date.today()
        filename    = self.output_dir / f"{report_date:%Y-%m-%d}_OptionVolumeLeaders.xlsx"

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        self._write_contracts_sheet(wb, "Stocks",     stock_contracts)
        self._write_bets_sheet(     wb, "Stocks Bets", stock_biases)
        self._write_contracts_sheet(wb, "ETFs",        etf_contracts)
        self._write_bets_sheet(     wb, "ETFs Bets",   etf_biases)

        wb.save(filename)
        logger.info("Report saved → %s", filename)
        return filename

    # -----------------------------------------------------------------------
    # Sheet 1 / 3: raw contracts
    # -----------------------------------------------------------------------

    _CONTRACT_COLS = [
        ("Ticker",          12),
        ("Type",             6),
        ("Strike",           9),
        ("Expiration",      12),
        ("DTE",              6),
        ("Bid",              8),
        ("Ask",              8),
        ("Last",             8),
        ("Volume",          10),
        ("Open Interest",   14),
        ("Vol/OI Ratio",    12),
        ("Premium ($)",     14),
        ("Volatility",      11),
        ("Underlying $",    12),
        ("Trade Time",      18),
        ("Symbol",          28),
    ]

    def _write_contracts_sheet(
        self, wb: openpyxl.Workbook, name: str, contracts: list[OptionContract]
    ) -> None:
        ws = wb.create_sheet(name)
        headers = [c[0] for c in self._CONTRACT_COLS]
        self._write_header(ws, headers)

        for c in contracts:
            row = [
                c.ticker,
                c.option_type,
                c.strike,
                c.expiration_label,
                c.dte,
                c.bid,
                c.ask,
                c.last,
                c.volume,
                c.open_interest,
                c.vol_oi_ratio,
                round(c.premium_usd, 0),
                f"{c.volatility:.1%}",
                c.underlying_price,
                str(c.trade_time.strftime("%Y-%m-%d %H:%M")),
                c.symbol,
            ]
            ws.append(row)
            # Colour code by type
            fill = _GREEN if c.is_call else _RED
            for cell in ws[ws.max_row]:
                cell.fill = fill

        self._set_col_widths(ws, self._CONTRACT_COLS)
        ws.freeze_panes = "A2"

    # -----------------------------------------------------------------------
    # Sheet 2 / 4: Bets (aggregated bias per ticker)
    # -----------------------------------------------------------------------

    _BETS_COLS = [
        ("Ticker",             10),
        ("Bias",                8),
        ("Signal Strength",    15),
        ("Consecutive Days",   16),
        ("Parity (Vol)",       13),
        ("Parity (Premium)",   15),
        ("Total Premium ($)",  16),
        ("Call Volume",        12),
        ("Put Volume",         12),
        ("Call Premium ($)",   14),
        ("Put Premium ($)",    14),
        ("Total Contracts",    15),
        ("OI Change",          11),
        ("Volume Change",      14),
        # Dominant contract columns
        ("Top Strike",         10),
        ("Top Type",            8),
        ("Top Expiry",         10),
        ("Top DTE",             8),
        ("Top Premium ($)",    14),
        ("Top Vol/OI",         10),
        ("Top IV",              8),
    ]

    def _write_bets_sheet(
        self, wb: openpyxl.Workbook, name: str, biases: list[TickerBias]
    ) -> None:
        ws = wb.create_sheet(name)
        headers = [c[0] for c in self._BETS_COLS]
        self._write_header(ws, headers)

        for b in biases:
            dom = b.dominant_contract
            row = [
                b.ticker,
                b.bias,
                b.signal_strength,
                b.consecutive_days,
                b.parity,
                b.premium_parity,
                round(b.total_premium, 0),
                b.call_volume,
                b.put_volume,
                round(b.call_premium, 0),
                round(b.put_premium, 0),
                b.total_contracts,
                b.oi_change,
                b.volume_change,
                # Dominant contract
                dom.strike           if dom else "",
                dom.option_type      if dom else "",
                dom.expiration_label if dom else "",
                dom.dte              if dom else "",
                round(dom.premium_usd, 0) if dom else "",
                dom.vol_oi_ratio     if dom else "",
                f"{dom.volatility:.1%}" if dom else "",
            ]
            ws.append(row)

            # Row colour: Strong + Long/Short gets gold highlight
            if b.signal_strength == "Strong":
                fill = _GOLD
            elif b.bias == "Long":
                fill = _GREEN
            elif b.bias == "Short":
                fill = _RED
            else:
                continue  # neutral — no fill

            for cell in ws[ws.max_row]:
                cell.fill = fill

        self._set_col_widths(ws, self._BETS_COLS)
        ws.freeze_panes = "A2"

    # -----------------------------------------------------------------------
    # Helpers
    # -----------------------------------------------------------------------

    @staticmethod
    def _write_header(ws, headers: list[str]) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.fill      = _HEADER
            cell.font      = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _set_col_widths(ws, col_defs: list[tuple]) -> None:
        for i, (_, width) in enumerate(col_defs, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
